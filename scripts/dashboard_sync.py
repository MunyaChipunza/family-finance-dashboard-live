from __future__ import annotations

import argparse
import datetime as dt
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
BUNDLE_DIR = SCRIPT_DIR.parent
DEFAULT_WORKBOOK = (BUNDLE_DIR.parent / "Finance.xlsx").resolve()
DEFAULT_OUTPUT = (BUNDLE_DIR / "dashboard_data.json").resolve()
DEFAULT_LOCAL_SNAPSHOT = (BUNDLE_DIR.parent / "Finance_Executive_Dashboard.html").resolve()
DEFAULT_STATE = (SCRIPT_DIR / ".sync_state.json").resolve()
TIMEZONE = dt.timezone(dt.timedelta(hours=2), name="SAST")
HEADER_SENTINEL = ("Status", "Section", "Group", "Item")
MONTHLY_SECTIONS = ("Income", "Monthly Cost", "Debt Payment", "Savings Contribution")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Refresh finance dashboard JSON from the live workbook.")
    parser.add_argument("--workbook", help="Optional workbook path.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Optional output JSON path.")
    return parser.parse_args()


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace(",", "").replace("R", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def iso_datetime(value: dt.datetime | None) -> str | None:
    if value is None:
        return None
    return value.astimezone(TIMEZONE).isoformat()


def fmt_money(value: float | None, currency: str = "ZAR") -> str:
    if value is None:
        return "-"
    symbol = "$" if currency.upper() == "USD" else "R"
    return f"{symbol} {value:,.2f}"


def pct(value: float | None) -> str:
    if value is None:
        return "-"
    return f"{value * 100:.1f}%"


def clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(value, maximum))


def load_workbook_source(workbook: str | None) -> tuple[Path, Any]:
    path = Path(workbook).expanduser().resolve() if workbook else DEFAULT_WORKBOOK
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    try:
        return path, load_workbook(path, data_only=False)
    except Exception as exc:
        raise RuntimeError(
            f"Could not read workbook directly: {path}. Save and close Excel if the file is mid-write, then try again."
        ) from exc


def row_values(ws, row_number: int) -> list[Any]:
    return [cell.value for cell in ws[row_number]]


def find_header_row(ws) -> int:
    limit = min(ws.max_row, 60)
    for row_number in range(1, limit + 1):
        values = tuple(clean_text(value) for value in row_values(ws, row_number)[:4])
        if values == HEADER_SENTINEL:
            return row_number
    raise ValueError("Could not find the finance data header row.")


def workbook_rows(ws, header_row: int) -> list[dict[str, Any]]:
    headers = [clean_text(value) for value in row_values(ws, header_row)]
    records: list[dict[str, Any]] = []
    blank_streak = 0
    for row_number in range(header_row + 1, ws.max_row + 1):
        values = row_values(ws, row_number)
        if not any(value not in (None, "") for value in values[: len(headers)]):
            blank_streak += 1
            if blank_streak >= 12:
                break
            continue
        blank_streak = 0
        record = {
            headers[index]: values[index] if index < len(values) else None for index in range(len(headers))
        }
        if clean_text(record.get("Item")):
            records.append(record)
    return records


def include_row(row: dict[str, Any]) -> bool:
    status = clean_text(row.get("Status")).lower()
    return status not in {"closed", "remove", "archive", "ignore"}


def section_name(row: dict[str, Any]) -> str:
    return clean_text(row.get("Section"))


def tag_name(row: dict[str, Any]) -> str:
    return clean_text(row.get("Dashboard Tag")).lower()


def monthly_amount(row: dict[str, Any], field: str) -> float:
    value = parse_float(row.get(field))
    return value if value is not None else 0.0


def balance_amount(row: dict[str, Any]) -> float:
    value = parse_float(row.get("Current Balance"))
    return value if value is not None else 0.0


def rows_for_section(rows: list[dict[str, Any]], name: str) -> list[dict[str, Any]]:
    return [row for row in rows if section_name(row).lower() == name.lower()]


def sum_monthly(rows: list[dict[str, Any]], section: str, field: str) -> float:
    return sum(monthly_amount(row, field) for row in rows_for_section(rows, section))


def sum_balances(rows: list[dict[str, Any]], section: str, tags: set[str] | None = None) -> float:
    total = 0.0
    for row in rows_for_section(rows, section):
        if tags is not None and tag_name(row) not in tags:
            continue
        total += balance_amount(row)
    return total


def line_item(row: dict[str, Any]) -> dict[str, Any]:
    budget = monthly_amount(row, "Budget Monthly")
    actual = monthly_amount(row, "Actual This Month")
    return {
        "section": section_name(row),
        "group": clean_text(row.get("Group")),
        "item": clean_text(row.get("Item")),
        "owner": clean_text(row.get("Owner")),
        "budget": budget,
        "actual": actual,
        "variance": actual - budget,
        "currency": clean_text(row.get("Currency")) or "ZAR",
        "timing": clean_text(row.get("Timing")),
        "auto": clean_text(row.get("Auto")),
        "tag": clean_text(row.get("Dashboard Tag")),
        "priority": clean_text(row.get("Priority")),
        "notes": clean_text(row.get("Notes")),
    }


def balance_item(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "section": section_name(row),
        "group": clean_text(row.get("Group")),
        "item": clean_text(row.get("Item")),
        "owner": clean_text(row.get("Owner")),
        "balance": balance_amount(row),
        "currency": clean_text(row.get("Currency")) or "ZAR",
        "timing": clean_text(row.get("Timing")),
        "tag": clean_text(row.get("Dashboard Tag")),
        "priority": clean_text(row.get("Priority")),
        "notes": clean_text(row.get("Notes")),
    }


def monthly_rows(rows: list[dict[str, Any]], sections: tuple[str, ...]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows:
        if section_name(row) in sections:
            result.append(line_item(row))
    return result


def balance_rows(rows: list[dict[str, Any]], section: str, tags: set[str] | None = None) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows_for_section(rows, section):
        if tags is not None and tag_name(row) not in tags:
            continue
        result.append(balance_item(row))
    result.sort(key=lambda item: item["balance"], reverse=True)
    return result


def actuals_captured(rows: list[dict[str, Any]]) -> bool:
    has_any_actual = False
    has_variance_from_budget = False
    for row in monthly_rows(rows, MONTHLY_SECTIONS):
        if abs(row["actual"]) > 0.009:
            has_any_actual = True
        if abs((row["actual"] or 0.0) - (row["budget"] or 0.0)) > 0.009:
            has_variance_from_budget = True
    return has_any_actual and has_variance_from_budget


def normalize_actuals(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], bool]:
    if actuals_captured(rows):
        return rows, False
    normalized: list[dict[str, Any]] = []
    for row in rows:
        copy = dict(row)
        if section_name(copy) in MONTHLY_SECTIONS:
            copy["Actual This Month"] = monthly_amount(copy, "Budget Monthly")
        normalized.append(copy)
    return normalized, True


def build_scorecard(
    income_actual: float,
    surplus_actual: float,
    runway_months: float | None,
    debt_service_ratio: float | None,
    savings_rate: float | None,
    net_worth: float,
) -> tuple[int, list[dict[str, Any]]]:
    surplus_ratio = (surplus_actual / income_actual) if income_actual else 0.0
    liquidity = round(clamp((runway_months or 0.0) / 6.0, 0.0, 1.0) * 35)
    margin = round(clamp(surplus_ratio / 0.20, 0.0, 1.0) * 25)
    leverage = round(clamp((0.35 - (debt_service_ratio or 0.0)) / 0.35, 0.0, 1.0) * 20)
    discipline = round(clamp((savings_rate or 0.0) / 0.20, 0.0, 1.0) * 10)
    balance_sheet = 10 if net_worth > 0 else 0
    score = int(liquidity + margin + leverage + discipline + balance_sheet)
    pillars = [
        {"label": "Liquidity", "score": liquidity, "outOf": 35},
        {"label": "Margin", "score": margin, "outOf": 25},
        {"label": "Leverage", "score": leverage, "outOf": 20},
        {"label": "Discipline", "score": discipline, "outOf": 10},
        {"label": "Balance Sheet", "score": balance_sheet, "outOf": 10},
    ]
    return score, pillars


def health_signal(score: int) -> dict[str, str]:
    if score >= 80:
        return {"label": "Strong", "tone": "good"}
    if score >= 60:
        return {"label": "Stable", "tone": "info"}
    if score >= 40:
        return {"label": "Tight", "tone": "warn"}
    return {"label": "Critical", "tone": "bad"}


def build_executive_summary(
    report_month: str,
    signal: dict[str, str],
    actual_surplus: float,
    liquid_cash: float,
    reserves: float,
    debt_service_ratio: float | None,
    runway_months: float | None,
    net_worth: float,
    top_cost_rows: list[dict[str, Any]],
    using_budget_as_actuals: bool,
) -> str:
    direction = "surplus" if actual_surplus >= 0 else "deficit"
    runway_label = f"{runway_months:.1f} months" if runway_months is not None else "not yet available"
    top_names = ", ".join(row["item"] for row in top_cost_rows[:3]) if top_cost_rows else "no major lines yet"
    source_note = "Budget values are temporarily standing in as actuals. " if using_budget_as_actuals else ""
    return (
        f"{report_month}: household health reads {signal['label'].lower()}. "
        f"{source_note}The current operating view shows a monthly {direction} of {fmt_money(abs(actual_surplus))}, "
        f"with {fmt_money(liquid_cash)} in liquid cash, {fmt_money(reserves)} in reserves, "
        f"debt service at {pct(debt_service_ratio)}, runway at {runway_label}, and tracked net worth at {fmt_money(net_worth)}. "
        f"The largest active monthly drivers are {top_names}."
    )


def build_focus_items(
    actual_surplus: float,
    runway_months: float | None,
    debt_service_ratio: float | None,
    savings_rate: float | None,
    watchlist_rows: list[dict[str, Any]],
    top_cost_rows: list[dict[str, Any]],
) -> list[dict[str, str]]:
    items: list[dict[str, str]] = []
    if actual_surplus < 0:
        items.append(
            {
                "title": "Close the monthly gap",
                "detail": f"Current actuals imply a deficit of {fmt_money(abs(actual_surplus))} per month.",
                "tone": "bad",
            }
        )
    if (runway_months or 0) < 3:
        items.append(
            {
                "title": "Strengthen reserves",
                "detail": f"Liquid cash plus reserves cover about {runway_months or 0:.1f} months of core outflows.",
                "tone": "warn",
            }
        )
    if (debt_service_ratio or 0) > 0.25:
        items.append(
            {
                "title": "Debt load is still heavy",
                "detail": f"Debt payments are consuming {pct(debt_service_ratio)} of take-home income.",
                "tone": "warn" if (debt_service_ratio or 0) <= 0.35 else "bad",
            }
        )
    if (savings_rate or 0) == 0:
        items.append(
            {
                "title": "Turn surplus into deliberate savings",
                "detail": "Use the planned savings rows to track reserve or investment transfers as soon as they happen.",
                "tone": "info",
            }
        )
    if watchlist_rows:
        items.append(
            {
                "title": "Watch the pressure points",
                "detail": f"Priority lines right now: {', '.join(row['item'] for row in watchlist_rows[:3])}.",
                "tone": "warn",
            }
        )
    if not items and top_cost_rows:
        items.append(
            {
                "title": "Operating model is under control",
                "detail": f"Biggest monthly lines remain {', '.join(row['item'] for row in top_cost_rows[:3])}.",
                "tone": "good",
            }
        )
    return items[:4]


def performance_row(label: str, budget: float, actual: float, income_line: bool = False) -> dict[str, Any]:
    variance = actual - budget
    if label == "Net Position":
        tone = "good" if actual >= 0 else "bad"
    elif income_line:
        tone = "good" if variance >= 0 else "warn"
    else:
        tone = "bad" if variance > 0.009 else "good" if variance < -0.009 else "info"
    return {
        "label": label,
        "budget": budget,
        "actual": actual,
        "variance": variance,
        "tone": tone,
        "ratio": (actual / budget) if budget else None,
    }


def build_performance_rows(
    income_budget: float,
    income_actual: float,
    costs_budget: float,
    costs_actual: float,
    debt_budget: float,
    debt_actual: float,
    savings_budget: float,
    savings_actual: float,
) -> list[dict[str, Any]]:
    outflows_budget = costs_budget + debt_budget + savings_budget
    outflows_actual = costs_actual + debt_actual + savings_actual
    return [
        performance_row("Income", income_budget, income_actual, income_line=True),
        performance_row("Core Costs", costs_budget, costs_actual),
        performance_row("Debt Service", debt_budget, debt_actual),
        performance_row("Savings Allocation", savings_budget, savings_actual),
        performance_row("Net Position", income_budget - outflows_budget, income_actual - outflows_actual, income_line=True),
    ]


def build_capital_stack(
    liquid_cash: float,
    reserves: float,
    medical_saver: float,
    investments: float,
    retirement: float,
    working_float: float,
    debt_balances: float,
    net_worth: float,
) -> list[dict[str, Any]]:
    return [
        {"label": "Liquid Cash", "amount": liquid_cash, "tone": "info", "detail": "Current transaction accounts"},
        {"label": "Emergency Reserves", "amount": reserves, "tone": "good", "detail": "Notice savings set aside"},
        {"label": "Medical Saver", "amount": medical_saver, "tone": "info", "detail": "Momentum HealthSaver balance"},
        {"label": "Working Float", "amount": working_float, "tone": "info", "detail": "Usable card float currently tracked"},
        {"label": "Investments", "amount": investments, "tone": "good", "detail": "Tax-free and brokerage capital"},
        {"label": "Retirement", "amount": retirement, "tone": "good", "detail": "Long-term retirement assets"},
        {"label": "Liabilities", "amount": -debt_balances, "tone": "bad", "detail": "Tracked vehicle debt balances"},
        {"label": "Net Worth", "amount": net_worth, "tone": "good" if net_worth >= 0 else "bad", "detail": "Tracked assets less tracked liabilities"},
    ]


def top_monthly_lines(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates = monthly_rows(rows, ("Monthly Cost", "Debt Payment"))
    candidates.sort(key=lambda row: row["actual"] if row["actual"] else row["budget"], reverse=True)
    return candidates[:8]


def over_budget_lines(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    candidates = monthly_rows(rows, ("Monthly Cost", "Debt Payment", "Savings Contribution"))
    flagged = [row for row in candidates if row["budget"] > 0 and row["actual"] > row["budget"]]
    flagged.sort(key=lambda row: row["variance"], reverse=True)
    return flagged


def build_watchlist(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    flagged = over_budget_lines(rows)
    if flagged:
        return [
            {
                "item": row["item"],
                "owner": row["owner"],
                "budget": row["budget"],
                "actual": row["actual"],
                "variance": row["variance"],
                "reason": "Over budget",
                "tone": "bad",
            }
            for row in flagged[:8]
        ]

    candidates = [row for row in top_monthly_lines(rows) if row["priority"] in {"Critical", "High"}]
    return [
        {
            "item": row["item"],
            "owner": row["owner"],
            "budget": row["budget"],
            "actual": row["actual"],
            "variance": row["variance"],
            "reason": "Core exposure",
            "tone": "warn",
        }
        for row in candidates[:6]
    ]


def build_debt_highlights(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    liabilities = balance_rows(rows, "Liability Balance")
    debt_rows = {row["item"].lower(): row for row in monthly_rows(rows, ("Debt Payment",))}
    result = []
    for item in liabilities:
        debt_line = debt_rows.get(item["item"].lower())
        instalment = 0.0
        if debt_line:
            instalment = debt_line["actual"] or debt_line["budget"]
        result.append(
            {
                "item": item["item"],
                "owner": item["owner"],
                "balance": item["balance"],
                "instalment": instalment,
                "timing": debt_line["timing"] if debt_line else item["timing"],
                "notes": debt_line["notes"] if debt_line and debt_line["notes"] else item["notes"],
            }
        )
    return result


def build_open_items(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in rows_for_section(rows, "Open Item"):
        result.append(
            {
                "area": clean_text(row.get("Group")) or "Open Item",
                "question": clean_text(row.get("Item")),
                "assumption": clean_text(row.get("Timing")),
                "update": clean_text(row.get("Notes")),
            }
        )
    return result


def render_local_snapshot(payload: dict[str, Any]) -> str:
    data_blob = json.dumps(payload)
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Family Finance Executive Dashboard</title>
<style>
:root{{--bg:#f6efe7;--panel:#fff8f0;--line:#d9c3ad;--ink:#2f2219;--muted:#7e6957;--emerald:#6a4a35;--gold:#b08662;--coral:#8f5f40;--aqua:#c09b79;--shadow:0 22px 50px rgba(72,49,33,.14)}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{min-height:100vh;background:linear-gradient(180deg,#f9f1e8,#ecdfcf 58%,#f5e9db);color:var(--ink);font:14px/1.5 'Segoe UI',Arial,sans-serif}}
.shell{{max-width:1480px;margin:0 auto;padding:28px 20px 40px}}
.panel,.card,.surface,.table-shell,.note{{background:linear-gradient(180deg,rgba(255,252,246,.98),rgba(255,247,238,.98));border:1px solid rgba(106,74,53,.16);border-radius:24px;box-shadow:var(--shadow)}}
.hero{{padding:28px;display:grid;grid-template-columns:1.1fr .9fr;gap:22px;background:linear-gradient(135deg,rgba(255,250,244,.98),rgba(244,230,214,.98) 54%,rgba(231,214,195,.98))}}
.eyebrow{{font:11px Consolas,monospace;letter-spacing:1.8px;text-transform:uppercase;color:#86624a;margin-bottom:10px}}
.title{{font:700 48px/1 Georgia,serif}}
.title span{{color:var(--coral)}}
.subtitle{{margin-top:10px;max-width:760px;color:#5a4637;font-size:16px}}
.exec{{margin-top:18px;font-size:15px;color:#4d392b;max-width:800px}}
.focus-grid{{margin-top:18px;display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px}}
.focus{{padding:16px;border-radius:18px;border:1px solid rgba(106,74,53,.10);background:rgba(255,252,247,.7)}}
.focus.good{{background:rgba(166,128,96,.14)}} .focus.warn{{background:rgba(196,163,131,.16)}} .focus.bad{{background:rgba(143,95,64,.12)}} .focus.info{{background:rgba(192,155,121,.16)}}
.focus h4{{font-size:16px}} .focus p{{margin-top:8px;color:#6a5849}}
.status{{display:inline-flex;align-items:center;gap:10px;padding:8px 14px;border-radius:999px;border:1px solid rgba(106,74,53,.20);background:rgba(106,74,53,.10);font:600 11px Consolas,monospace;text-transform:uppercase;letter-spacing:1px;color:var(--emerald)}}
.status.warn{{border-color:rgba(176,134,98,.30);background:rgba(176,134,98,.14);color:#805b3b}} .status.bad{{border-color:rgba(182,94,67,.28);background:rgba(182,94,67,.12);color:var(--coral)}}
.meta{{font:11px/1.8 Consolas,monospace;color:#725e4d;text-align:right}}
.meta em{{font-style:normal;color:var(--emerald)}}
.score-shell{{display:grid;grid-template-columns:200px 1fr;gap:18px;align-items:center;padding:18px;border-radius:22px;background:rgba(255,250,244,.72);border:1px solid rgba(106,74,53,.10)}}
.score-ring{{width:170px;height:170px;border-radius:50%;display:grid;place-items:center;background:conic-gradient(var(--emerald) 0deg,rgba(106,74,53,.12) 0deg);padding:16px;margin:0 auto}}
.score-core{{width:100%;height:100%;border-radius:50%;background:linear-gradient(180deg,#fffdf8,#f3e7d9);display:grid;place-items:center;text-align:center;border:1px solid rgba(106,74,53,.10)}}
.score-number{{font:700 42px/1 Georgia,serif}} .score-label{{margin-top:6px;font:12px Consolas,monospace;color:#705846;text-transform:uppercase}}
.pillars{{display:grid;gap:12px}} .pillar-name{{font:11px Consolas,monospace;letter-spacing:1px;text-transform:uppercase;color:var(--muted)}}
.pillar-top{{display:flex;justify-content:space-between;gap:10px;align-items:center}} .pillar-bar{{margin-top:6px;height:10px;border-radius:999px;background:#e9dac8;overflow:hidden}} .pillar-fill{{height:100%;border-radius:999px;background:linear-gradient(90deg,var(--emerald),var(--aqua))}}
.section{{margin-top:22px}} .label{{font:11px Consolas,monospace;letter-spacing:1.4px;text-transform:uppercase;color:var(--emerald);margin-bottom:12px}}
.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:14px}} .card{{padding:16px;border-top:4px solid rgba(106,74,53,.18)}} .card.good{{border-top-color:var(--emerald)}} .card.warn{{border-top-color:var(--gold)}} .card.bad{{border-top-color:var(--coral)}} .card.info{{border-top-color:var(--aqua)}}
.card-k{{font:10px Consolas,monospace;letter-spacing:1px;text-transform:uppercase;color:var(--muted)}} .card-v{{margin-top:8px;font-size:30px;font-weight:700;line-height:1}} .card-d{{margin-top:8px;color:#6c5848;font-size:12px}}
.split{{display:grid;grid-template-columns:1fr 1fr;gap:18px}} .surface{{padding:20px}}
.bar-list{{display:grid;gap:12px}} .bar-row{{padding:14px 16px;border-radius:18px;background:rgba(255,252,247,.7);border:1px solid rgba(106,74,53,.10)}}
.bar-top{{display:flex;justify-content:space-between;gap:14px;align-items:flex-start}} .bar-title{{font-size:16px;font-weight:700}} .bar-meta{{margin-top:5px;color:#6e5948;font-size:12px}} .bar-value{{font:700 18px Consolas,monospace;white-space:nowrap}}
.track{{margin-top:12px;height:10px;border-radius:999px;background:#ead9c7;overflow:hidden}} .fill{{height:100%;border-radius:999px;background:linear-gradient(90deg,var(--emerald),var(--aqua))}} .fill.warn{{background:linear-gradient(90deg,var(--gold),#caa17e)}} .fill.bad{{background:linear-gradient(90deg,var(--coral),#b88461)}}
.table-shell{{padding:18px}} .table-wrap{{overflow:auto;border:1px solid rgba(106,74,53,.10);border-radius:18px}}
table{{width:100%;border-collapse:collapse;min-width:760px}} th,td{{padding:12px 14px;border-bottom:1px solid rgba(106,74,53,.08);text-align:left}}
th{{background:#f2e4d3;color:#725c4a;font:11px Consolas,monospace;text-transform:uppercase;letter-spacing:1px}} td{{color:#2f241d}} tr:nth-child(even) td{{background:rgba(235,226,211,.22)}}
.pill{{display:inline-block;padding:4px 10px;border-radius:999px;font:11px Consolas,monospace;background:rgba(106,74,53,.08);color:var(--emerald)}} .note-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:14px}} .note{{padding:16px}} .note h4{{font-size:16px}} .note p{{margin-top:8px;color:#6f5b4b}}
@media (max-width:980px){{.hero,.split,.score-shell{{grid-template-columns:1fr}} .meta{{text-align:left}}}}
</style>
</head>
<body>
<div class="shell">
  <header class="hero panel">
    <div>
      <div class="eyebrow">Local Executive Snapshot</div>
      <div class="title">Family Finance <span>Executive Dashboard</span></div>
      <p class="subtitle" id="subtitle"></p>
      <p class="exec" id="executiveSummary"></p>
      <div class="focus-grid" id="focusItems"></div>
    </div>
    <div>
      <div style="display:flex;justify-content:space-between;gap:14px;align-items:flex-start;flex-wrap:wrap">
        <div class="status" id="statusBadge"><span id="statusText"></span></div>
        <div class="meta">
          <div>Mode: <em id="dataMode"></em></div>
          <div>Report month: <em id="reportMonth"></em></div>
          <div>Workbook: <em id="sourceName"></em></div>
          <div>Generated: <em id="generatedAt"></em></div>
        </div>
      </div>
      <div class="score-shell" style="margin-top:16px">
        <div class="score-ring" id="scoreRing"><div class="score-core"><div><div class="score-number" id="scoreValue"></div><div class="score-label" id="scoreLabel"></div></div></div></div>
        <div class="pillars" id="pillarScores"></div>
      </div>
    </div>
  </header>
  <section class="section"><div class="label">Board Summary</div><div class="cards" id="summaryCards"></div></section>
  <section class="section split"><div class="surface"><div class="label">Operating Model</div><div class="bar-list" id="performanceRows"></div></div><div class="surface"><div class="label">Capital Stack</div><div class="bar-list" id="capitalStack"></div></div></section>
  <section class="section split"><div class="surface"><div class="label">Top Monthly Cost Drivers</div><div class="bar-list" id="topCosts"></div></div><div class="surface"><div class="label">Watchlist</div><div class="bar-list" id="watchlist"></div></div></section>
  <section class="section split"><div class="table-shell"><div class="label">Cash, Reserves, And Float</div><div class="table-wrap"><table id="cashTable"></table></div></div><div class="table-shell"><div class="label">Debt Profile</div><div class="table-wrap"><table id="debtTable"></table></div></div></section>
  <section class="section"><div class="label">Open Items</div><div class="note-grid" id="openItems"></div></section>
</div>
<script>
const dashboard = {data_blob};
function esc(v){{return String(v ?? '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#39;')}}
function fmtMoney(v,c='ZAR'){{if(v===null||v===undefined||v==='') return '-'; if(String(c).toUpperCase()==='USD') return new Intl.NumberFormat('en-US',{{style:'currency',currency:'USD',maximumFractionDigits:2}}).format(Number(v)); return new Intl.NumberFormat('en-ZA',{{style:'currency',currency:'ZAR',maximumFractionDigits:2}}).format(Number(v));}}
function fmtDateTime(v){{if(!v) return '-'; try{{return new Intl.DateTimeFormat('en-ZA',{{day:'2-digit',month:'short',year:'numeric',hour:'2-digit',minute:'2-digit',second:'2-digit',hour12:false,timeZone:'Africa/Johannesburg'}}).format(new Date(v))}}catch{{return v}}}}
function toneClass(t){{return ['good','warn','bad','info'].includes(t) ? t : 'info'}}
function renderCards(){{const host=document.getElementById('summaryCards');host.innerHTML=(dashboard.summaryCards||[]).map(item=>`<div class="card ${{toneClass(item.tone)}}"><div class="card-k">${{esc(item.label)}}</div><div class="card-v">${{esc(item.value)}}</div><div class="card-d">${{esc(item.detail)}}</div></div>`).join('')}}
function renderFocus(){{const host=document.getElementById('focusItems');host.innerHTML=(dashboard.focusItems||[]).map(item=>`<div class="focus ${{toneClass(item.tone)}}"><h4>${{esc(item.title)}}</h4><p>${{esc(item.detail)}}</p></div>`).join('')}}
function renderScore(){{const score=Number(dashboard.healthScore||0);document.getElementById('scoreValue').textContent=score;document.getElementById('scoreLabel').textContent=`${{dashboard.health?.label||'Health'}} Score`;document.getElementById('scoreRing').style.background=`conic-gradient(var(--emerald) ${{Math.max(0,Math.min(100,score))*3.6}}deg, rgba(106,74,53,.12) 0deg)`;document.getElementById('pillarScores').innerHTML=(dashboard.pillarScores||[]).map(item=>{{const width=item.outOf?Math.max(0,Math.min(100,(Number(item.score)/Number(item.outOf))*100)):0;return `<div><div class="pillar-top"><div class="pillar-name">${{esc(item.label)}}</div><div class="pill">${{esc(String(item.score))}}/${{esc(String(item.outOf))}}</div></div><div class="pillar-bar"><div class="pillar-fill" style="width:${{width}}%"></div></div></div>`}}).join('')}}
function renderBarList(id,items,valueFn,titleFn,metaFn,displayFn,toneFn){{const host=document.getElementById(id);if(!(items||[]).length){{host.innerHTML='<div class="bar-row"><div class="bar-title">No data yet</div></div>';return;}}const maxValue=Math.max(...items.map(item=>Math.abs(Number(valueFn(item)||0))),1);host.innerHTML=items.map(item=>{{const value=Number(valueFn(item)||0);const width=Math.max(6,Math.min(100,(Math.abs(value)/maxValue)*100));const tone=toneClass(toneFn(item));return `<div class="bar-row"><div class="bar-top"><div><div class="bar-title">${{esc(titleFn(item))}}</div><div class="bar-meta">${{esc(metaFn(item))}}</div></div><div class="bar-value">${{esc(displayFn(item))}}</div></div><div class="track"><div class="fill ${{tone}}" style="width:${{width}}%"></div></div></div>`}}).join('')}}
function renderTable(id,headers,rows){{const table=document.getElementById(id);if(!(rows||[]).length){{table.innerHTML='<tbody><tr><td>No rows available.</td></tr></tbody>';return;}}table.innerHTML=`<thead><tr>${{headers.map(h=>`<th>${{esc(h.label)}}</th>`).join('')}}</tr></thead><tbody>${{rows.map(row=>`<tr>${{headers.map(h=>`<td>${{h.render?h.render(row):esc(row[h.key] ?? '')}}</td>`).join('')}}</tr>`).join('')}}</tbody>`}}
function renderOpen(){{const host=document.getElementById('openItems');host.innerHTML=(dashboard.openItems||[]).map(item=>`<div class="note"><h4>${{esc(item.area)}}</h4><p>${{esc(item.question)}}</p><p><span class="pill">Current</span> ${{esc(item.assumption||'-')}}</p><p><span class="pill">Update</span> ${{esc(item.update||'-')}}</p></div>`).join('')}}
document.getElementById('subtitle').textContent=dashboard.subtitle||'';document.getElementById('executiveSummary').textContent=dashboard.executiveSummary||'';document.getElementById('dataMode').textContent=dashboard.dataMode||'-';document.getElementById('reportMonth').textContent=dashboard.reportMonth||'-';document.getElementById('sourceName').textContent=dashboard.sourceName||'-';document.getElementById('generatedAt').textContent=fmtDateTime(dashboard.generatedAt);document.getElementById('statusBadge').className=`status ${{toneClass(dashboard.health?.tone)}}`;document.getElementById('statusText').textContent=dashboard.health?.label||'Live';
renderFocus();renderScore();renderCards();
renderBarList('performanceRows',dashboard.performanceRows||[],item=>item.actual,item=>item.label,item=>`Budget ${{fmtMoney(item.budget)}} | Actual ${{fmtMoney(item.actual)}} | Variance ${{fmtMoney(item.variance)}}`,item=>fmtMoney(item.actual),item=>item.tone);
renderBarList('capitalStack',dashboard.capitalStack||[],item=>item.amount,item=>item.label,item=>item.detail||'',item=>fmtMoney(item.amount),item=>item.tone);
renderBarList('topCosts',dashboard.topCostRows||[],item=>item.actual||item.budget,item=>item.item,item=>`${{item.group}} | ${{item.owner||'Household'}}`,item=>fmtMoney(item.actual||item.budget),()=> 'warn');
renderBarList('watchlist',dashboard.watchlistRows||[],item=>Math.abs(item.variance||item.actual||item.budget),item=>item.item,item=>`${{item.reason}} | Budget ${{fmtMoney(item.budget)}} | Actual ${{fmtMoney(item.actual)}}`,item=>item.reason==='Over budget'?fmtMoney(item.variance):fmtMoney(item.actual||item.budget),item=>item.tone||'warn');
renderTable('cashTable',[{{label:'Account',key:'item'}},{{label:'Owner',key:'owner'}},{{label:'Type',key:'tag'}},{{label:'Balance',render:r=>fmtMoney(r.balance,r.currency)}},{{label:'Notes',key:'notes'}}],dashboard.cashAccounts||[]);
renderTable('debtTable',[{{label:'Debt',key:'item'}},{{label:'Outstanding',render:r=>fmtMoney(r.balance)}},{{label:'Monthly',render:r=>fmtMoney(r.instalment)}},{{label:'Timing',key:'timing'}},{{label:'Notes',key:'notes'}}],dashboard.debtHighlights||[]);
renderOpen();
</script>
</body>
</html>"""


def write_local_snapshot(payload: dict[str, Any], target: Path = DEFAULT_LOCAL_SNAPSHOT) -> Path:
    target.write_text(render_local_snapshot(payload), encoding="utf-8")
    return target


def refresh_dashboard_data(workbook: str | None = None, output: str | Path = DEFAULT_OUTPUT) -> Path:
    workbook_path, wb = load_workbook_source(workbook)
    ws = wb[wb.sheetnames[0]]
    header_row = find_header_row(ws)
    raw_rows = [row for row in workbook_rows(ws, header_row) if include_row(row)]
    rows, using_budget_as_actuals = normalize_actuals(raw_rows)

    income_budget = sum_monthly(rows, "Income", "Budget Monthly")
    income_actual = sum_monthly(rows, "Income", "Actual This Month")
    costs_budget = sum_monthly(rows, "Monthly Cost", "Budget Monthly")
    costs_actual = sum_monthly(rows, "Monthly Cost", "Actual This Month")
    debt_budget = sum_monthly(rows, "Debt Payment", "Budget Monthly")
    debt_actual = sum_monthly(rows, "Debt Payment", "Actual This Month")
    savings_budget = sum_monthly(rows, "Savings Contribution", "Budget Monthly")
    savings_actual = sum_monthly(rows, "Savings Contribution", "Actual This Month")

    outflows_budget = costs_budget + debt_budget + savings_budget
    outflows_actual = costs_actual + debt_actual + savings_actual
    surplus_budget = income_budget - outflows_budget
    surplus_actual = income_actual - outflows_actual

    liquid_cash = sum_balances(rows, "Asset Balance", {"cash"})
    reserves = sum_balances(rows, "Asset Balance", {"reserve"})
    medical_saver = sum_balances(rows, "Asset Balance", {"medical"})
    investments = sum_balances(rows, "Asset Balance", {"investment"})
    retirement = sum_balances(rows, "Asset Balance", {"retirement"})
    working_float = sum_balances(rows, "Asset Balance", {"working-float"})
    debt_balances = sum_balances(rows, "Liability Balance")
    net_worth = liquid_cash + reserves + medical_saver + investments + retirement + working_float - debt_balances

    debt_service_ratio = (debt_actual / income_actual) if income_actual else None
    savings_rate = (savings_actual / income_actual) if income_actual else None
    core_outflows = costs_actual + debt_actual
    runway_months = ((liquid_cash + reserves) / core_outflows) if core_outflows else None
    coverage_ratio = (income_actual / outflows_actual) if outflows_actual else None

    score, pillar_scores = build_scorecard(
        income_actual=income_actual,
        surplus_actual=surplus_actual,
        runway_months=runway_months,
        debt_service_ratio=debt_service_ratio,
        savings_rate=savings_rate,
        net_worth=net_worth,
    )
    signal = health_signal(score)

    operating_lines = monthly_rows(rows, ("Monthly Cost", "Debt Payment", "Savings Contribution"))
    operating_lines.sort(key=lambda row: (row["section"], -(row["actual"] if row["actual"] else row["budget"])))
    top_cost_rows = top_monthly_lines(rows)
    watchlist_rows = build_watchlist(rows)
    report_month = clean_text(ws["B5"].value) or dt.datetime.now(TIMEZONE).strftime("%B %Y")

    summary_cards = [
        {"label": "Actual Income", "value": fmt_money(income_actual), "detail": "Current month income run-rate", "tone": "good"},
        {"label": "Actual Outflows", "value": fmt_money(outflows_actual), "detail": "Costs, debt, and savings allocations", "tone": "warn"},
        {
            "label": "Monthly Surplus",
            "value": fmt_money(surplus_actual),
            "detail": "Income minus outflows",
            "tone": "good" if surplus_actual >= 0 else "bad",
        },
        {
            "label": "Runway",
            "value": f"{runway_months:.1f} months" if runway_months is not None else "-",
            "detail": "Liquid cash plus reserves vs core outflows",
            "tone": "good" if (runway_months or 0) >= 3 else "warn" if (runway_months or 0) >= 1 else "bad",
        },
        {
            "label": "Debt Service",
            "value": pct(debt_service_ratio),
            "detail": "Debt payments as share of take-home",
            "tone": "good" if (debt_service_ratio or 0) <= 0.20 else "warn" if (debt_service_ratio or 0) <= 0.35 else "bad",
        },
        {
            "label": "Savings Rate",
            "value": pct(savings_rate),
            "detail": "Formal savings as share of income",
            "tone": "good" if (savings_rate or 0) >= 0.15 else "warn" if (savings_rate or 0) > 0 else "bad",
        },
        {"label": "Liquid Cash", "value": fmt_money(liquid_cash), "detail": "Current accounts only", "tone": "info"},
        {"label": "Net Worth", "value": fmt_money(net_worth), "detail": "Tracked assets less tracked liabilities", "tone": "good" if net_worth >= 0 else "bad"},
    ]

    payload = {
        "title": "Family Finance Command Deck",
        "subtitle": "Executive view of household cash, commitments, debt, and long-term capital.",
        "reportMonth": report_month,
        "sourceName": workbook_path.name,
        "generatedAt": iso_datetime(dt.datetime.now(TIMEZONE)),
        "sourceUpdatedAt": iso_datetime(dt.datetime.fromtimestamp(workbook_path.stat().st_mtime, tz=TIMEZONE)),
        "refreshSeconds": 60,
        "dataMode": "Budget-backed actuals" if using_budget_as_actuals else "Live actuals",
        "health": signal,
        "healthScore": score,
        "pillarScores": pillar_scores,
        "executiveSummary": build_executive_summary(
            report_month=report_month,
            signal=signal,
            actual_surplus=surplus_actual,
            liquid_cash=liquid_cash,
            reserves=reserves,
            debt_service_ratio=debt_service_ratio,
            runway_months=runway_months,
            net_worth=net_worth,
            top_cost_rows=top_cost_rows,
            using_budget_as_actuals=using_budget_as_actuals,
        ),
        "focusItems": build_focus_items(
            actual_surplus=surplus_actual,
            runway_months=runway_months,
            debt_service_ratio=debt_service_ratio,
            savings_rate=savings_rate,
            watchlist_rows=watchlist_rows,
            top_cost_rows=top_cost_rows,
        ),
        "summaryCards": summary_cards,
        "performanceRows": build_performance_rows(
            income_budget=income_budget,
            income_actual=income_actual,
            costs_budget=costs_budget,
            costs_actual=costs_actual,
            debt_budget=debt_budget,
            debt_actual=debt_actual,
            savings_budget=savings_budget,
            savings_actual=savings_actual,
        ),
        "capitalStack": build_capital_stack(
            liquid_cash=liquid_cash,
            reserves=reserves,
            medical_saver=medical_saver,
            investments=investments,
            retirement=retirement,
            working_float=working_float,
            debt_balances=debt_balances,
            net_worth=net_worth,
        ),
        "topCostRows": top_cost_rows,
        "watchlistRows": watchlist_rows,
        "cashAccounts": balance_rows(rows, "Asset Balance", {"cash", "reserve", "medical", "working-float"}),
        "investmentRows": balance_rows(rows, "Asset Balance", {"investment", "retirement"}),
        "debtHighlights": build_debt_highlights(rows),
        "operatingLines": operating_lines,
        "incomeRows": monthly_rows(rows, ("Income",)),
        "openItems": build_open_items(rows),
        "totals": {
            "budgetIncome": income_budget,
            "actualIncome": income_actual,
            "budgetOutflows": outflows_budget,
            "actualOutflows": outflows_actual,
            "budgetSurplus": surplus_budget,
            "actualSurplus": surplus_actual,
            "liquidCash": liquid_cash,
            "reserves": reserves,
            "medicalSaver": medical_saver,
            "investments": investments,
            "retirement": retirement,
            "workingFloat": working_float,
            "debtBalances": debt_balances,
            "coverageRatio": coverage_ratio,
        },
    }

    output_path = Path(output).expanduser()
    if not output_path.is_absolute():
        output_path = (BUNDLE_DIR / output_path).resolve()
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    write_local_snapshot(payload)
    return output_path


def main() -> None:
    args = parse_args()
    output_path = refresh_dashboard_data(workbook=args.workbook, output=args.output)
    print(output_path)


if __name__ == "__main__":
    main()
