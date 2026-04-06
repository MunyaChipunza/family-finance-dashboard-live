from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).resolve().parent
BUNDLE_DIR = SCRIPT_DIR.parent
DEFAULT_WORKBOOK = (BUNDLE_DIR.parent / "Finance_Input.xlsx").resolve()
DEFAULT_OUTPUT = (BUNDLE_DIR / "dashboard_data.json").resolve()
TIMEZONE = dt.timezone(dt.timedelta(hours=2), name="SAST")
CREATE_NO_WINDOW = getattr(subprocess, "CREATE_NO_WINDOW", 0)
SECTION_NAMES = {
    "INCOME",
    "BANK / CASH",
    "CARD / CREDIT",
    "ACTIVE DEBTS",
    "MONTHLY COSTS",
    "INVESTMENTS / RETIREMENT",
    "REMOVED / EXCLUDED",
    "OPEN ITEMS",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build finance dashboard JSON from the Finance_Input workbook.")
    parser.add_argument("--workbook", help="Local workbook path.")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output JSON path.")
    return parser.parse_args()


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace(",", "").replace("R", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def iso_datetime(value: dt.datetime | None) -> str | None:
    if value is None:
        return None
    return value.astimezone(TIMEZONE).isoformat()


def fmt_money(value: float | None, currency: str = "ZAR") -> str:
    if value is None:
        return "-"
    symbol = "$" if currency == "USD" else "R"
    return f"{symbol} {value:,.2f}"


def create_snapshot(source_path: Path) -> Path:
    if os.name != "nt":
        return source_path

    helper = SCRIPT_DIR / "save_excel_snapshot.ps1"
    temp_dir = Path(tempfile.mkdtemp(prefix="finance_workbook_"))
    target = temp_dir / source_path.name
    command = [
        "powershell",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(helper),
        "-SourcePath",
        str(source_path),
        "-TargetPath",
        str(target),
    ]
    startupinfo = subprocess.STARTUPINFO()
    startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    startupinfo.wShowWindow = 0
    result = subprocess.run(command, text=True, capture_output=True, creationflags=CREATE_NO_WINDOW, startupinfo=startupinfo)
    if result.returncode == 0 and target.exists():
        return target
    return source_path


def load_workbook_snapshot(workbook: str | None):
    path = Path(workbook).expanduser().resolve() if workbook else DEFAULT_WORKBOOK
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    snapshot = create_snapshot(path)
    return path, load_workbook(snapshot, data_only=False)


def row_values(ws, row_number: int) -> list[Any]:
    values = [cell.value for cell in ws[row_number]]
    while values and values[-1] in (None, ""):
        values.pop()
    return values


def read_sections(ws) -> dict[str, list[dict[str, Any]]]:
    sections: dict[str, list[dict[str, Any]]] = {}
    row = 1
    while row <= ws.max_row:
        name = clean_text(ws.cell(row, 1).value)
        if name not in SECTION_NAMES:
            row += 1
            continue
        headers = [clean_text(value) for value in row_values(ws, row + 1)]
        data_rows: list[dict[str, Any]] = []
        current = row + 2
        while current <= ws.max_row:
            values = row_values(ws, current)
            if not values:
                break
            if clean_text(values[0]) in SECTION_NAMES:
                break
            item = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
            data_rows.append(item)
            current += 1
        sections[name] = data_rows
        row = current + 1
    return sections


def include_row(row: dict[str, Any]) -> bool:
    review = clean_text(row.get("Review")).lower()
    return review != "remove"


def build_income_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "owner": clean_text(row.get("Owner")),
            "source": clean_text(row.get("Income Source")),
            "monthlyMin": parse_float(row.get("Net Monthly Min")),
            "monthlyMax": parse_float(row.get("Net Monthly Max")),
            "notes": clean_text(row.get("Notes")),
        }
        for row in rows
        if include_row(row)
    ]


def build_cash_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "owner": clean_text(row.get("Owner")),
            "institution": clean_text(row.get("Institution")),
            "product": clean_text(row.get("Account / Product")),
            "accountType": clean_text(row.get("Account Type")),
            "balance": parse_float(row.get("Balance")),
            "currency": "ZAR",
            "asOf": clean_text(row.get("As Of")),
            "treatment": clean_text(row.get("Treatment")),
            "notes": clean_text(row.get("Notes")),
        }
        for row in rows
        if include_row(row)
    ]


def build_card_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "owner": clean_text(row.get("Owner")),
            "institution": clean_text(row.get("Institution")),
            "product": clean_text(row.get("Product")),
            "currentBalance": parse_float(row.get("Current Balance")),
            "pendingTransactions": parse_float(row.get("Pending Transactions")),
            "availableBalance": parse_float(row.get("Available Balance")),
            "creditLimit": parse_float(row.get("Credit Limit")),
            "paymentDue": clean_text(row.get("Payment Due")),
            "minPayment": parse_float(row.get("Min Payment")),
            "effectiveRate": clean_text(row.get("Effective Rate")),
            "treatment": clean_text(row.get("Treatment")),
            "notes": clean_text(row.get("Notes")),
        }
        for row in rows
        if include_row(row)
    ]


def build_debt_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "owner": clean_text(row.get("Owner")),
            "lender": clean_text(row.get("Lender")),
            "product": clean_text(row.get("Debt / Product")),
            "monthlyInstalment": parse_float(row.get("Monthly Instalment")),
            "observedDebit": parse_float(row.get("Observed Bank Debit")),
            "outstandingBalance": parse_float(row.get("Outstanding Balance")),
            "interestRate": clean_text(row.get("Interest Rate")),
            "debitTiming": clean_text(row.get("Debit Timing")),
            "status": clean_text(row.get("Status")),
            "notes": clean_text(row.get("Notes")),
        }
        for row in rows
        if include_row(row)
    ]


def build_cost_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items = [
        {
            "category": clean_text(row.get("Category")),
            "item": clean_text(row.get("Item")),
            "monthlyMin": parse_float(row.get("Monthly Amount Min")),
            "monthlyMax": parse_float(row.get("Monthly Amount Max")),
            "treatment": clean_text(row.get("Treatment")),
            "notes": clean_text(row.get("Notes")),
        }
        for row in rows
        if include_row(row)
    ]
    items.sort(key=lambda item: item["monthlyMax"] or item["monthlyMin"] or 0, reverse=True)
    return items


def build_investment_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "owner": clean_text(row.get("Owner")),
            "institution": clean_text(row.get("Institution")),
            "product": clean_text(row.get("Product")),
            "balance": parse_float(row.get("Balance")),
            "currency": clean_text(row.get("Currency")) or "ZAR",
            "asOf": clean_text(row.get("As Of")),
            "treatment": clean_text(row.get("Treatment")),
            "notes": clean_text(row.get("Notes")),
        }
        for row in rows
        if include_row(row)
    ]


def build_excluded_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "item": clean_text(row.get("Item")),
            "reason": clean_text(row.get("Why Excluded / Removed")),
            "referenceAmount": parse_float(row.get("Reference Amount")),
            "notes": clean_text(row.get("Notes")),
        }
        for row in rows
        if include_row(row)
    ]


def build_open_items(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "area": clean_text(row.get("Area")),
            "question": clean_text(row.get("Question / Missing Detail")),
            "assumption": clean_text(row.get("Current Assumption")),
            "update": clean_text(row.get("Your Update")),
        }
        for row in rows
        if any(clean_text(value) for value in row.values())
    ]


def total(values: list[float | None]) -> float:
    return sum(value or 0.0 for value in values)


def build_summary_cards(income_rows, cash_rows, debt_rows, cost_rows, investment_rows) -> list[dict[str, str]]:
    income_min = total([row["monthlyMin"] for row in income_rows if row["owner"] != "Household"])
    income_max = total([row["monthlyMax"] or row["monthlyMin"] for row in income_rows if row["owner"] != "Household"])
    cost_min = total([row["monthlyMin"] for row in cost_rows])
    cost_max = total([row["monthlyMax"] or row["monthlyMin"] for row in cost_rows])
    debt_instalments = total([row["monthlyInstalment"] for row in debt_rows if row["status"].lower() == "active"])
    tracked_min = cost_min + debt_instalments
    tracked_max = cost_max + debt_instalments
    liquid_cash = total([row["balance"] for row in cash_rows if "cash" in row["treatment"].lower() or "day-to-day" in row["treatment"].lower()])
    reserve_savings = total([row["balance"] for row in cash_rows if ("reserve" in row["treatment"].lower() or "savings" in row["treatment"].lower()) and "medical" not in row["treatment"].lower()])
    medical_saver = total([row["balance"] for row in cash_rows if "medical" in row["treatment"].lower()])
    investments_total = total([row["balance"] for row in investment_rows if row["treatment"] in {"Summary only", "Retirement asset"}])
    debt_balance = total([row["outstandingBalance"] for row in debt_rows if row["status"].lower() == "active"])
    return [
        {"label": "Net Income", "value": f"{fmt_money(income_min)} - {fmt_money(income_max)}", "detail": "Combined monthly take-home range", "tone": "good"},
        {"label": "Monthly Costs", "value": f"{fmt_money(cost_min)} - {fmt_money(cost_max)}", "detail": "Confirmed monthly commitments only", "tone": "warn"},
        {"label": "Debt Instalments", "value": fmt_money(debt_instalments), "detail": "Active monthly debt repayments", "tone": "bad"},
        {"label": "Tracked Outflows", "value": f"{fmt_money(tracked_min)} - {fmt_money(tracked_max)}", "detail": "Monthly costs plus debt instalments", "tone": "warn"},
        {"label": "Liquid Cash", "value": fmt_money(liquid_cash), "detail": "Current accounts only", "tone": "good"},
        {"label": "Reserves", "value": fmt_money(reserve_savings), "detail": "Notice and reserve savings", "tone": "info"},
        {"label": "Medical Saver", "value": fmt_money(medical_saver), "detail": "Momentum HealthSaver balance", "tone": "info"},
        {"label": "Investments", "value": fmt_money(investments_total), "detail": "Long-term investment and retirement totals", "tone": "good"},
        {"label": "Debt Balances", "value": fmt_money(debt_balance), "detail": "Outstanding active debt balances", "tone": "bad"},
    ]


def build_cash_highlights(cash_rows) -> list[dict[str, Any]]:
    return [
        {
            "label": f"{row['institution']} - {row['product']}",
            "meta": f"{row['owner']} | {row['treatment']}",
            "amount": row["balance"],
            "currency": row["currency"],
        }
        for row in cash_rows
    ]


def build_debt_highlights(card_rows, debt_rows) -> list[dict[str, Any]]:
    items = [
        {
            "label": f"{row['institution']} - {row['product']}",
            "meta": f"{row['owner']} | {row['treatment']}",
            "amount": row["currentBalance"],
        }
        for row in card_rows
    ]
    items.extend(
        {
            "label": row["product"],
            "meta": f"{row['lender']} | {row['debitTiming']}",
            "balance": row["outstandingBalance"],
            "instalment": row["monthlyInstalment"],
        }
        for row in debt_rows
    )
    return items


def refresh_dashboard_data(workbook: str | None = None, output: str | Path = DEFAULT_OUTPUT) -> Path:
    workbook_path, wb = load_workbook_snapshot(workbook)
    ws = wb[wb.sheetnames[0]]
    sections = read_sections(ws)

    income_rows = build_income_rows(sections.get("INCOME", []))
    cash_rows = build_cash_rows(sections.get("BANK / CASH", []))
    card_rows = build_card_rows(sections.get("CARD / CREDIT", []))
    debt_rows = build_debt_rows(sections.get("ACTIVE DEBTS", []))
    cost_rows = build_cost_rows(sections.get("MONTHLY COSTS", []))
    investment_rows = build_investment_rows(sections.get("INVESTMENTS / RETIREMENT", []))
    excluded_rows = build_excluded_rows(sections.get("REMOVED / EXCLUDED", []))
    open_items = build_open_items(sections.get("OPEN ITEMS", []))

    payload = {
        "title": "Family Finance Dashboard",
        "subtitle": "Single-workbook dashboard for cash, commitments, debt, and long-term assets.",
        "sourceName": workbook_path.name,
        "generatedAt": iso_datetime(dt.datetime.now(TIMEZONE)),
        "sourceUpdatedAt": iso_datetime(dt.datetime.fromtimestamp(workbook_path.stat().st_mtime, tz=TIMEZONE)),
        "refreshSeconds": 60,
        "summaryCards": build_summary_cards(income_rows, cash_rows, debt_rows, cost_rows, investment_rows),
        "cashHighlights": build_cash_highlights(cash_rows),
        "debtHighlights": build_debt_highlights(card_rows, debt_rows),
        "incomeRows": income_rows,
        "cashRows": cash_rows,
        "cardRows": card_rows,
        "debtRows": debt_rows,
        "costRows": cost_rows,
        "investmentRows": investment_rows,
        "excludedRows": excluded_rows,
        "openItems": open_items,
    }

    output_path = Path(output).expanduser()
    if not output_path.is_absolute():
        output_path = (BUNDLE_DIR / output_path).resolve()
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return output_path


def main() -> None:
    args = parse_args()
    output_path = refresh_dashboard_data(workbook=args.workbook, output=args.output)
    print(output_path)


if __name__ == "__main__":
    main()

